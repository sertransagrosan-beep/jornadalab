import streamlit as st
import pandas as pd
import numpy as np
import io
import requests
import time
from functools import lru_cache
from typing import Optional, Tuple, Dict, List
import hashlib
import pickle
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# ==============================
# CONFIGURACIÓN DE PÁGINA
# ==============================
st.set_page_config(
    page_title="Jornada Laboral Conductores",
    page_icon="🚛",
    layout="wide"
)

# ==============================
# CACHÉ PERSISTENTE
# ==============================
CACHE_DIR = Path("cache")
CACHE_DIR.mkdir(exist_ok=True)

def get_cache_key(lat: float, lon: float) -> str:
    """Genera clave de caché para coordenadas"""
    return hashlib.md5(f"{round(lat,4)}_{round(lon,4)}".encode()).hexdigest()

def save_to_cache(key: str, value: str):
    """Guarda resultado en caché persistente"""
    cache_file = CACHE_DIR / f"{key}.pkl"
    with open(cache_file, 'wb') as f:
        pickle.dump(value, f)

def load_from_cache(key: str) -> Optional[str]:
    """Carga resultado del caché persistente"""
    cache_file = CACHE_DIR / f"{key}.pkl"
    if cache_file.exists():
        with open(cache_file, 'rb') as f:
            return pickle.load(f)
    return None

# ==============================
# FUNCIÓN PARA AUTO-AJUSTAR COLUMNAS EN EXCEL
# ==============================
def autoajustar_columnas(archivo_excel: io.BytesIO, nombre_hoja: str):
    """Autoajusta el ancho de las columnas en una hoja específica"""
    # Guardar el archivo temporalmente
    temp_path = Path("temp_ajuste.xlsx")
    with open(temp_path, 'wb') as f:
        f.write(archivo_excel.getvalue())
    
    # Cargar y ajustar
    wb = load_workbook(temp_path)
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # Calcular longitud del contenido
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = min(cell_length, 50)  # Limitar a 50 caracteres
                except:
                    pass
            
            # Ajustar ancho con un margen
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar cambios
    wb.save(temp_path)
    
    # Leer de vuelta al buffer
    with open(temp_path, 'rb') as f:
        archivo_excel.seek(0)
        archivo_excel.truncate(0)
        archivo_excel.write(f.read())
    
    # Limpiar archivo temporal
    temp_path.unlink()

# ==============================
# CONFIGURACIÓN EN SIDEBAR
# ==============================
with st.sidebar:
    st.header("⚙️ Configuración")
    
    HORAS_MAX_JORNADA = st.number_input(
        "Horas máximas jornada", 
        value=8.0, 
        min_value=1.0, 
        max_value=24.0,
        step=0.5,
        help="Duración máxima permitida de la jornada laboral"
    )
    
    HORAS_DESCANSO_LARGO = st.number_input(
        "Horas descanso largo", 
        value=4.0, 
        min_value=1.0, 
        max_value=12.0,
        step=0.5,
        help="Tiempo mínimo considerado como descanso largo"
    )
    
    col1, col2 = st.columns(2)
    with col1:
        MIN_PAUSA = st.number_input(
            "Pausa mínima (min)", 
            value=34, 
            min_value=5, 
            max_value=120,
            step=5
        )
    with col2:
        MIN_PARADA = st.number_input(
            "Parada mínima (min)", 
            value=17, 
            min_value=5, 
            max_value=60,
            step=5
        )
    
    RADIO_CLUSTER = st.slider(
        "Radio cluster (metros)", 
        min_value=50, 
        max_value=1000, 
        value=300, 
        step=50,
        help="Distancia máxima para agrupar ubicaciones similares"
    )
    
    st.divider()
    st.caption("🔧 Optimizaciones activadas:")
    st.caption("- ✅ Caché persistente de geolocalización")
    st.caption("- ✅ Procesamiento vectorizado")
    st.caption("- ✅ Lazy loading de datos")

# ==============================
# CONVERSIÓN DE UNIDADES
# ==============================
HORAS_MIN_PAUSA = MIN_PAUSA / 60
UMBRAL_PARADA_MIN = MIN_PARADA / 60

# ==============================
# 🌍 GEO (CON CACHÉ PERSISTENTE)
# ==============================
@st.cache_data(ttl=3600, max_entries=1000)
def coord_a_municipio(lat: float, lon: float) -> str:
    """Obtiene municipio desde coordenadas con caché persistente"""
    
    if pd.isna(lat) or pd.isna(lon):
        return ""
    
    cache_key = get_cache_key(lat, lon)
    
    # Intentar cargar de caché persistente
    cached_result = load_from_cache(cache_key)
    if cached_result:
        return cached_result
    
    try:
        # Intentar con caché de Streamlit primero
        url = "https://nominatim.openstreetmap.org/reverse"
        params = {
            "lat": lat, 
            "lon": lon, 
            "format": "json",
            "zoom": 10  # Reduce detalle para mejorar velocidad
        }
        headers = {"User-Agent": "StreamlitJornadaApp/1.0"}
        
        response = requests.get(url, params=params, headers=headers, timeout=3)
        
        if response.status_code == 200:
            data = response.json()
            address = data.get("address", {})
            
            # Priorizar niveles administrativos
            ciudad = (
                address.get("city") or
                address.get("town") or
                address.get("village") or
                address.get("municipality") or
                address.get("county") or
                ""
            )
            
            if ciudad:
                save_to_cache(cache_key, ciudad)
                time.sleep(0.5)  # Reducido de 1s a 0.5s
                return ciudad
                
    except requests.Timeout:
        st.warning(f"Timeout geocodificando ({lat}, {lon})")
    except Exception as e:
        st.warning(f"Error geocodificación: {str(e)[:50]}")
    
    # Fallback: coordenadas formateadas
    fallback = f"{round(lat,3)},{round(lon,3)}"
    save_to_cache(cache_key, fallback)
    return fallback

def obtener_localizacion(coordenadas: str) -> str:
    """Obtiene la localización formateada desde coordenadas"""
    if pd.isna(coordenadas) or coordenadas == "":
        return ""
    try:
        lat, lon = map(float, str(coordenadas).split(","))
        return coord_a_municipio(lat, lon)
    except:
        return ""

# ==============================
# LECTOR INTELIGENTE OPTIMIZADO
# ==============================
@st.cache_data(ttl=3600)
def leer_archivo(file) -> Optional[pd.DataFrame]:
    """Lee archivo CSV/Excel con optimizaciones"""
    try:
        if file.name.endswith(".xlsx"):
            # Leer solo columnas necesarias
            df = pd.read_excel(file, dtype_backend='numpy_nullable')
        else:
            # Optimizar lectura de CSV
            try:
                df = pd.read_csv(
                    file, 
                    sep=";", 
                    encoding="utf-8",
                    low_memory=False,
                    engine='c'  # Usar engine C para velocidad
                )
            except:
                file.seek(0)
                df = pd.read_csv(
                    file, 
                    sep=None, 
                    engine='python',
                    engine_kwargs={'nrows': 100000}  # Límite por seguridad
                )
        
        # Limpiar columnas solo una vez
        df.columns = df.columns.astype(str).str.strip()
        cols_to_drop = [col for col in df.columns if col.startswith('Unnamed')]
        if cols_to_drop:
            df = df.drop(columns=cols_to_drop)
        
        return df
    
    except Exception as e:
        st.error(f"Error leyendo {file.name}: {str(e)}")
        return None

# ==============================
# PROCESAMIENTO DE COORDENADAS VECTORIZADO
# ==============================
@st.cache_data
def parse_coords_vectorized(coord_series: pd.Series) -> pd.DataFrame:
    """Versión vectorizada para parsear coordenadas"""
    # Extraer latitud y longitud usando pandas vectorizado
    coords_df = coord_series.str.split(',', expand=True)
    
    if coords_df.shape[1] >= 2:
        lat = pd.to_numeric(coords_df[0], errors='coerce')
        lon = pd.to_numeric(coords_df[1], errors='coerce')
    else:
        lat = pd.Series([np.nan] * len(coord_series))
        lon = pd.Series([np.nan] * len(coord_series))
    
    return pd.DataFrame({'lat': lat, 'lon': lon})

def distancia_metros_vectorized(lat1, lon1, lat2, lon2):
    """Versión vectorizada del cálculo de distancia"""
    R = 6371000
    lat1_rad = np.radians(lat1)
    lat2_rad = np.radians(lat2)
    dlat = np.radians(lat2 - lat1)
    dlon = np.radians(lon2 - lon1)
    
    a = np.sin(dlat/2)**2 + np.cos(lat1_rad) * np.cos(lat2_rad) * np.sin(dlon/2)**2
    return 2 * R * np.arctan2(np.sqrt(a), np.sqrt(1 - a))

# ==============================
# CLUSTERING OPTIMIZADO
# ==============================
def clusterizar_ubicaciones(df: pd.DataFrame, radio: float = 300) -> List[Dict]:
    """Clustering optimizado con numpy"""
    if df.empty:
        return []
    
    clusters = []
    
    # Convertir a arrays numpy para velocidad
    puntos = df[['lat', 'lon']].values
    pesos = df['peso'].values
    
    for i, (lat, lon) in enumerate(puntos):
        if np.isnan(lat):
            continue
        
        mejor_cluster = None
        mejor_dist = radio
        
        # Buscar cluster cercano
        for j, cluster in enumerate(clusters):
            dist = distancia_metros_vectorized(
                lat, lon, cluster['lat'], cluster['lon']
            )
            if dist < mejor_dist:
                mejor_dist = dist
                mejor_cluster = j
        
        if mejor_cluster is not None:
            # Actualizar cluster existente
            cluster = clusters[mejor_cluster]
            peso_nuevo = cluster['peso'] + pesos[i]
            cluster['lat'] = (cluster['lat'] * cluster['peso'] + lat * pesos[i]) / peso_nuevo
            cluster['lon'] = (cluster['lon'] * cluster['peso'] + lon * pesos[i]) / peso_nuevo
            cluster['peso'] = peso_nuevo
            cluster['count'] += 1
        else:
            # Nuevo cluster
            clusters.append({
                'lat': lat,
                'lon': lon,
                'peso': pesos[i],
                'count': 1
            })
    
    return clusters

def obtener_ubic_principal(grupo: pd.DataFrame, radio: float = 300) -> str:
    """Obtiene ubicación principal del grupo"""
    
    if grupo.empty or 'Coordenadas' not in grupo.columns:
        return ""
    
    # Parsear coordenadas vectorizado
    coords_df = parse_coords_vectorized(grupo['Coordenadas'])
    grupo = grupo.assign(
        lat=coords_df['lat'],
        lon=coords_df['lon']
    )
    
    # Calcular pesos vectorizado
    grupo['peso'] = np.where(
        grupo['estado'].isin(['ralenti', 'apagado']),
        grupo['delta_horas'] * 2,
        grupo['delta_horas'] * 0.3
    )
    
    grupo_valid = grupo.dropna(subset=['lat'])
    
    if grupo_valid.empty:
        return ""
    
    clusters = clusterizar_ubicaciones(grupo_valid, radio)
    
    if not clusters:
        return ""
    
    mejor = max(clusters, key=lambda x: x['peso'])
    
    return coord_a_municipio(mejor['lat'], mejor['lon'])

# ==============================
# PROCESAMIENTO PRINCIPAL OPTIMIZADO
# ==============================
@st.cache_data(ttl=3600)
def procesar_datos(df_original: pd.DataFrame, config: Dict) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Procesamiento principal optimizado - retorna KPIs y Bloques"""
    
    df = df_original.copy()
    
    # Limpieza ASAP
    df['fecha_hora'] = pd.to_datetime(df['fecha_hora'], errors='coerce')
    df = df.dropna(subset=['fecha_hora']).reset_index(drop=True)
    
    # Procesar ignición vectorizado
    df['ignicion_on'] = df.get('ignicion', '').astype(str).str.lower().isin(['encendido', 'true', '1'])
    
    # Procesar velocidad vectorizado
    if 'velocidad' in df.columns:
        velocidad_str = df['velocidad'].astype(str)
        velocidad_num = velocidad_str.str.replace(',', '.', regex=False)
        df['velocidad'] = pd.to_numeric(velocidad_num, errors='coerce').fillna(0)
    else:
        df['velocidad'] = 0
    
    # Ordenar eficientemente
    df = df.sort_values(['vehiculo', 'fecha_hora']).reset_index(drop=True)
    
    # Calcular estados vectorizado - CORREGIDO
    df['estado'] = 'apagado'
    df.loc[df['ignicion_on'] & (df['velocidad'] > 0), 'estado'] = 'conduciendo'
    df.loc[df['ignicion_on'] & (df['velocidad'] == 0), 'estado'] = 'ralenti'
    
    # Calcular delta horas vectorizado - CORREGIDO
    df['fecha_siguiente'] = df.groupby('vehiculo')['fecha_hora'].shift(-1)
    df['delta_horas'] = (df['fecha_siguiente'] - df['fecha_hora']).dt.total_seconds() / 3600
    df['delta_horas'] = df['delta_horas'].fillna(0)
    
    # Agrupar cambios de estado
    df['grupo'] = (df['estado'] != df['estado'].shift()).cumsum()
    
    # Agregar por grupos
    bloques = df.groupby(['vehiculo', 'grupo']).agg({
        'estado': 'first',
        'fecha_hora': ['min', 'max'],
        'delta_horas': 'sum'
    }).reset_index()
    
    bloques.columns = ['vehiculo', 'grupo', 'estado', 'inicio', 'fin', 'duracion_horas']
    
    # Agregar ubicaciones a bloques
    # Obtener localización del inicio y fin de cada bloque
    bloques['inicio_localizacion'] = ''
    bloques['fin_localizacion'] = ''
    
    for idx, row in bloques.iterrows():
        # Buscar la localización en el momento del inicio
        inicio_data = df[df['fecha_hora'] == row['inicio']]
        if not inicio_data.empty and 'Localización' in inicio_data.columns:
            bloques.at[idx, 'inicio_localizacion'] = obtener_localizacion(inicio_data['Localización'].iloc[0])
        
        # Buscar la localización en el momento del fin
        fin_data = df[df['fecha_hora'] == row['fin']]
        if not fin_data.empty and 'Localización' in fin_data.columns:
            bloques.at[idx, 'fin_localizacion'] = obtener_localizacion(fin_data['Localización'].iloc[0])
    
    # KPIs optimizados
    kpis_list = []
    
    for (vehiculo, fecha), grupo in df.groupby(['vehiculo', df['fecha_hora'].dt.date]):
        
        # Obtener conductor (primero no nulo)
        conductores = grupo['conductor'].dropna()
        conductor = conductores.iloc[0] if not conductores.empty else "Desconocido"
        
        # Filtrar solo cuando ignición está encendida
        ignicion_on = grupo[grupo['ignicion_on']]
        if ignicion_on.empty:
            continue
        
        inicio_jornada = ignicion_on['fecha_hora'].min()
        fin_jornada = ignicion_on['fecha_hora'].max()
        
        # Calcular horas por estado - CORREGIDO: ahora sí calcula correctamente
        horas_conduccion = grupo.loc[grupo['estado'] == 'conduciendo', 'delta_horas'].sum()
        horas_ralenti = grupo.loc[grupo['estado'] == 'ralenti', 'delta_horas'].sum()
        horas_trabajo = horas_conduccion + horas_ralenti
        
        # Ubicación final
        ult_coords = grupo['Coordenadas'].dropna()
        if not ult_coords.empty:
            lat, lon = parse_coords_vectorized(pd.Series([ult_coords.iloc[-1]])).iloc[0]
            ubicacion = coord_a_municipio(lat, lon) if not pd.isna(lat) else ""
        else:
            ubicacion = ""
        
        # Obtener ubicación principal
        ubic_principal = obtener_ubic_principal(grupo, config['radio_cluster'])
        
        # Calcular paradas, descansos y pausas
        bloques_v = bloques[bloques['vehiculo'] == vehiculo]
        
        numero_paradas = 0
        horas_descanso = 0
        horas_pausa = 0
        
        fecha_ts = pd.Timestamp(fecha)
        next_day = fecha_ts + pd.Timedelta(days=1)
        
        for _, b in bloques_v.iterrows():
            inicio = max(b['inicio'], fecha_ts)
            fin = min(b['fin'], next_day)
            
            if inicio < fin:
                horas = (fin - inicio).total_seconds() / 3600
                
                if b['estado'] in ['ralenti', 'apagado'] and horas >= config['umbral_parada_min']:
                    numero_paradas += 1
                
                if b['estado'] == 'apagado':
                    if horas >= config['horas_descanso_largo']:
                        horas_descanso += horas
                    elif horas >= config['horas_min_pausa']:
                        horas_pausa += horas
        
        kpis_list.append({
            'conductor': conductor,
            'vehiculo': vehiculo,
            'fecha': fecha,
            'origen': '',
            'destino': '',
            'ubicación': ubicacion,
            'inicio_jornada': inicio_jornada,
            'fin_jornada': fin_jornada,
            'numero_paradas': numero_paradas,
            'horas_trabajo': round(horas_trabajo, 2),
            'horas_conduccion': round(horas_conduccion, 2),
            'horas_ralenti': round(horas_ralenti, 2),
            'horas_descanso': round(horas_descanso, 2),
            'horas_pausa': round(horas_pausa, 2),
            'ubic_principal': ubic_principal
        })
    
    if not kpis_list:
        return pd.DataFrame(), bloques
    
    kpis = pd.DataFrame(kpis_list)
    
    # Formatear horas
    kpis['inicio_jornada'] = pd.to_datetime(kpis['inicio_jornada']).dt.strftime('%I:%M %p').str.lstrip('0')
    kpis['fin_jornada'] = pd.to_datetime(kpis['fin_jornada']).dt.strftime('%I:%M %p').str.lstrip('0')
    
    return kpis, bloques

# ==============================
# FUNCIÓN PARA GENERAR EXCEL CON MÚLTIPLES HOJAS
# ==============================
def generar_excel_multiple(kpis_por_conductor: Dict, bloques_por_conductor: Dict, nombre_mes: str) -> io.BytesIO:
    """Genera Excel con múltiples hojas organizadas por conductor y vehículo"""
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Hoja de resumen general
        resumen_general = []
        for conductor, data in kpis_por_conductor.items():
            for _, row in data.iterrows():
                resumen_general.append({
                    'Conductor': conductor,
                    'Vehículo': row['vehiculo'],
                    'Fecha': row['fecha'],
                    'Horas Trabajo': row['horas_trabajo'],
                    'Horas Conducción': row['horas_conduccion'],
                    'Número Paradas': row['numero_paradas']
                })
        
        if resumen_general:
            df_resumen = pd.DataFrame(resumen_general)
            df_resumen.to_excel(writer, sheet_name='Resumen General', index=False)
        
        # Crear una hoja para cada combinación conductor-vehículo
        for conductor, kpis_df in kpis_por_conductor.items():
            for _, row in kpis_df.iterrows():
                vehiculo = row['vehiculo']
                fecha_jornada = row['fecha']
                
                # Nombre de la hoja para KPIs
                nombre_hoja_kpi = f"{conductor}_{vehiculo}".replace('/', '_').replace('\\', '_')[:31]
                
                # Crear DataFrame para este conductor y vehículo específico
                df_kpi = pd.DataFrame([row])
                df_kpi.to_excel(writer, sheet_name=nombre_hoja_kpi, index=False)
                
                # Hoja de bloques para este conductor-vehículo
                if conductor in bloques_por_conductor:
                    bloques_df = bloques_por_conductor[conductor]
                    bloques_vehiculo = bloques_df[bloques_df['vehiculo'] == vehiculo].copy()
                    
                    # Filtrar bloques por fecha
                    fecha_inicio = pd.Timestamp(fecha_jornada)
                    fecha_fin = fecha_inicio + pd.Timedelta(days=1)
                    
                    bloques_filtrados = bloques_vehiculo[
                        (bloques_vehiculo['inicio'] >= fecha_inicio) & 
                        (bloques_vehiculo['inicio'] < fecha_fin)
                    ]
                    
                    if not bloques_filtrados.empty:
                        nombre_hoja_bloques = f"Bloques_{conductor}_{vehiculo}".replace('/', '_').replace('\\', '_')[:31]
                        bloques_filtrados.to_excel(writer, sheet_name=nombre_hoja_bloques, index=False)
    
    # Autoajustar todas las columnas de todas las hojas
    for nombre_hoja in pd.ExcelWriter(output, engine='openpyxl').book.sheetnames:
        autoajustar_columnas(output, nombre_hoja)
    
    output.seek(0)
    return output

# ==============================
# INTERFAZ PRINCIPAL
# ==============================
st.title("🚛 Jornada Laboral Conductores")

# Subir archivos
files = st.file_uploader(
    "📂 Sube archivos CSV o Excel",
    accept_multiple_files=True,
    type=['csv', 'xlsx', 'xls'],
    help="Puedes subir múltiples archivos (diferentes conductores, vehículos y meses)"
)

if files:
    with st.spinner('📥 Cargando archivos...'):
        lista_df = []
        
        for file in files:
            df_temp = leer_archivo(file)
            if df_temp is not None and not df_temp.empty:
                # Mapeo flexible de columnas
                column_mapping = {
                    "Fecha y Hora": "fecha_hora",
                    "Velocidad": "velocidad", 
                    "Ignicion*": "ignicion",
                    "Conductor": "conductor",
                    "Localización": "Localización"
                }
                
                for old_name, new_name in column_mapping.items():
                    if old_name in df_temp.columns:
                        df_temp = df_temp.rename(columns={old_name: new_name})
                
                df_temp["vehiculo"] = file.name[:6].upper()
                lista_df.append(df_temp)
    
    if len(lista_df) == 0:
        st.error("❌ No se encontraron datos válidos en los archivos")
        st.stop()
    
    df = pd.concat(lista_df, ignore_index=True)
    
    with st.spinner('🔄 Procesando datos...'):
        # Configuración para procesamiento
        config = {
            'radio_cluster': RADIO_CLUSTER,
            'umbral_parada_min': UMBRAL_PARADA_MIN,
            'horas_descanso_largo': HORAS_DESCANSO_LARGO,
            'horas_min_pausa': HORAS_MIN_PAUSA,
            'horas_max_jornada': HORAS_MAX_JORNADA
        }
        
        kpis, bloques = procesar_datos(df, config)
    
    if kpis.empty:
        st.warning("⚠️ No se generaron KPIs. Verifica que los datos contengan información válida.")
        st.stop()
    
    # Mostrar resultados
    st.success(f"✅ Procesados {len(kpis)} registros")
    
    # Métricas rápidas
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total conductores", kpis['conductor'].nunique())
    with col2:
        st.metric("Total vehículos", kpis['vehiculo'].nunique())
    with col3:
        st.metric("Horas trabajo promedio", f"{kpis['horas_trabajo'].mean():.1f}h")
    with col4:
        st.metric("Horas conducción promedio", f"{kpis['horas_conduccion'].mean():.1f}h")
    
    # Dataframe interactivo
    st.subheader("📊 Resumen de Jornadas")
    
    # Filtros
    col1, col2 = st.columns(2)
    with col1:
        conductor_filter = st.multiselect(
            "Filtrar por conductor",
            options=kpis['conductor'].unique()
        )
    with col2:
        vehiculo_filter = st.multiselect(
            "Filtrar por vehículo", 
            options=kpis['vehiculo'].unique()
        )
    
    df_filtrado = kpis.copy()
    if conductor_filter:
        df_filtrado = df_filtrado[df_filtrado['conductor'].isin(conductor_filter)]
    if vehiculo_filter:
        df_filtrado = df_filtrado[df_filtrado['vehiculo'].isin(vehiculo_filter)]
    
    st.dataframe(
        df_filtrado,
        use_container_width=True,
        column_config={
            "inicio_jornada": st.column_config.TextColumn("Inicio jornada"),
            "fin_jornada": st.column_config.TextColumn("Fin jornada"),
            "horas_trabajo": st.column_config.NumberColumn("Horas trabajo", format="%.1f h"),
            "horas_conduccion": st.column_config.NumberColumn("Horas conducción", format="%.1f h"),
        }
    )
    
    # Mostrar bloques (opcional)
    with st.expander("📋 Ver detalles de bloques"):
        st.dataframe(bloques, use_container_width=True)
    
    # Exportar resultados
    st.subheader("💾 Exportar Resultados")
    
    # Preparar datos por conductor para exportación múltiple
    kpis_por_conductor = {}
    bloques_por_conductor = {}
    
    for conductor in kpis['conductor'].unique():
        kpis_por_conductor[conductor] = kpis[kpis['conductor'] == conductor]
        bloques_por_conductor[conductor] = bloques[bloques['vehiculo'].isin(
            kpis_por_conductor[conductor]['vehiculo'].unique()
        )]
    
    # Obtener el mes de los datos
    meses = kpis['fecha'].astype(str).str[:7].unique()
    nombre_mes = meses[0] if len(meses) > 0 else datetime.now().strftime("%Y-%m")
    
    # Generar nombre de archivo
    conductores_str = "_".join(kpis['conductor'].unique()[:3])
    if len(kpis['conductor'].unique()) > 3:
        conductores_str += "_y_otros"
    
    nombre_archivo = f"Jornada_Laboral_{conductores_str}_{nombre_mes}.xlsx"
    
    col1, col2 = st.columns(2)
    with col1:
        # Generar Excel con múltiples hojas
        excel_multiple = generar_excel_multiple(kpis_por_conductor, bloques_por_conductor, nombre_mes)
        
        st.download_button(
            label="📥 Descargar Excel (Múltiples hojas)",
            data=excel_multiple,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with col2:
        # Opción de descarga simple CSV
        csv = kpis.to_csv(index=False)
        st.download_button(
            label="📄 Descargar CSV (Resumen)",
            data=csv,
            file_name=f"Resumen_Jornada_{nombre_mes}.csv",
            mime="text/csv"
        )
    
    # Mostrar estadísticas adicionales
    with st.expander("📈 Estadísticas Avanzadas"):
        col1, col2 = st.columns(2)
        with col1:
            st.bar_chart(kpis.groupby('conductor')['horas_trabajo'].mean())
        with col2:
            st.bar_chart(kpis.groupby('vehiculo')['horas_trabajo'].mean())
        
        # Mostrar alertas si hay horas conducción en 0 (para debugging)
        if (kpis['horas_conduccion'] == 0).any():
            st.warning("⚠️ Algunos registros muestran 0 horas de conducción. Verifica que los datos contengan velocidad > 0 con ignición encendida.")

else:
    st.info("👈 Sube archivos CSV o Excel para comenzar el análisis")
    
    # Ejemplo de formato
    with st.expander("📋 Formato esperado de archivos"):
        st.markdown("""
        El archivo debe contener las siguientes columnas:
        - **Fecha y Hora**: timestamp de la lectura
        - **Velocidad**: velocidad del vehículo (km/h)
        - **Ignicion***: estado del encendido (encendido/apagado)
        - **Coordenadas**: latitud,longitud (opcional)
        - **Conductor**: identificador del conductor
        - **Localización**: texto con la ubicación (opcional)
        
        Puedes subir múltiples archivos CSV (separados por ;) o Excel.
        """)

# Limpiar caché si es necesario
if st.sidebar.button("🗑️ Limpiar caché"):
    st.cache_data.clear()
    for cache_file in CACHE_DIR.glob("*.pkl"):
        cache_file.unlink()
    st.success("Caché limpiada correctamente")
    st.rerun()
